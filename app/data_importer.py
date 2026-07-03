"""Importação inteligente unificada — CSV, Excel, SQL, SQLite."""
import csv
import io
import re
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from app.config import get_db_path as config_db_path
from app import database as db
from app.csv_io import (
    TABLE_CONFIG, auto_detect_table, read_csv_file,
    _detect_mapping, _normalize_header,
)

SUPPORTED_EXT = {".csv", ".xlsx", ".xls", ".sql", ".db", ".sqlite", ".sqlite3"}


def _map_rows(headers: list[str], rows: list[dict], table_key: str) -> tuple[list[str], list[dict]] | None:
    cfg = TABLE_CONFIG[table_key]
    mapping = _detect_mapping(headers, cfg["map"])
    if len(mapping) < 2:
        return None
    mapped: list[dict[str, Any]] = []
    for raw in rows:
        item: dict[str, Any] = {}
        for target, source in mapping.items():
            val = raw.get(source, "")
            if target == "dias_desde_ultima_compra":
                try:
                    item[target] = int(float(str(val).replace(",", "."))) if val else None
                except ValueError:
                    item[target] = None
            elif target in ("quantidade", "valor"):
                try:
                    item[target] = float(str(val).replace(",", ".").replace("€", "").strip()) if val else 0
                except ValueError:
                    item[target] = 0
            else:
                item[target] = str(val).strip() if val is not None else ""
        mapped.append(item)
    return list(mapping.keys()), mapped


def _import_table_data(table_key: str, headers: list[str], rows: list[dict],
                       db_path: Path | None, replace: bool) -> dict:
    mapped = _map_rows(headers, rows, table_key)
    if not mapped:
        return {"ok": False, "table": table_key, "imported": 0, "error": "Colunas nao reconhecidas"}
    columns, data = mapped
    if not data:
        return {"ok": False, "table": table_key, "imported": 0, "error": "Sem dados"}
    if replace:
        db.clear_table(TABLE_CONFIG[table_key]["table"], db_path)
    count = db.insert_rows(TABLE_CONFIG[table_key]["table"], columns, data, db_path)
    return {"ok": True, "table": table_key, "imported": count}


def import_csv_auto(path: Path, db_path: Path | None = None, replace: bool = True) -> dict:
    headers, rows = read_csv_file(path)
    if not rows:
        return {"ok": False, "error": "CSV vazio", "details": []}
    table = auto_detect_table(headers)
    if table == "clientes" and "vendas" in path.stem.lower():
        table = "vendas"
    elif "email" in path.stem.lower():
        table = "emails"
    result = _import_table_data(table, headers, rows, db_path, replace)
    return {"ok": result["ok"], "details": [result], "error": result.get("error", "")}


def import_excel(path: Path, db_path: Path | None = None) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl nao instalado", "details": []}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    details = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
        rows = []
        for row in rows_iter:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
            if any(d.values()):
                rows.append(d)
        if not rows:
            continue
        table = auto_detect_table(headers)
        sn = sheet_name.lower()
        if "venda" in sn:
            table = "vendas"
        elif "email" in sn:
            table = "emails"
        elif "client" in sn:
            table = "clientes"
        r = _import_table_data(table, headers, rows, db_path, replace=len(details) == 0)
        details.append(r)
    wb.close()
    ok = any(d.get("ok") for d in details)
    total = sum(d.get("imported", 0) for d in details)
    return {"ok": ok, "details": details, "total": total, "error": "" if ok else "Nenhuma folha reconhecida"}


def import_sqlite_file(path: Path, db_path: Path | None = None) -> dict:
    target = db_path or config_db_path()
    db.init_db(target)
    details = []
    src = sqlite3.connect(str(path))
    src.row_factory = sqlite3.Row
    dst = sqlite3.connect(str(target))
    for table in ("clientes", "vendas", "emails"):
        try:
            rows = src.execute(f"SELECT * FROM {table}").fetchall()
            if not rows:
                continue
            dst.execute(f"DELETE FROM {table}")
            cols = rows[0].keys()
            col_list = [c for c in cols if c != "id"]
            placeholders = ", ".join("?" for _ in col_list)
            sql = f"INSERT INTO {table} ({', '.join(col_list)}) VALUES ({placeholders})"
            for row in rows:
                dst.execute(sql, tuple(row[c] for c in col_list))
            details.append({"ok": True, "table": table, "imported": len(rows)})
        except sqlite3.OperationalError:
            pass
    dst.commit()
    src.close()
    dst.close()
    ok = any(d.get("ok") for d in details)
    return {"ok": ok, "details": details, "total": sum(d.get("imported", 0) for d in details)}


def import_sql_file(path: Path, db_path: Path | None = None) -> dict:
    target = db_path or config_db_path()
    db.init_db(target)
    text = path.read_text(encoding="utf-8", errors="replace")
    conn = sqlite3.connect(str(target))
    try:
        conn.executescript(text)
        conn.commit()
    except sqlite3.Error as e:
        conn.close()
        return {"ok": False, "error": str(e), "details": []}
    conn.close()
    details = []
    for table in ("clientes", "vendas", "emails"):
        try:
            n = db.count_table(table, target)
            if n > 0:
                details.append({"ok": True, "table": table, "imported": n})
        except Exception:
            pass
    return {"ok": bool(details), "details": details, "total": sum(d.get("imported", 0) for d in details)}


def import_any(path: Path, db_path: Path | None = None) -> dict:
    """Importa qualquer formato suportado — deteta automaticamente."""
    path = Path(path)
    if path.is_dir():
        results = []
        for f in sorted(path.iterdir()):
            if f.suffix.lower() in SUPPORTED_EXT:
                r = import_any(f, db_path)
                results.append(r)
        if not results:
            return {"ok": False, "error": "Pasta sem ficheiros suportados", "details": []}
        all_details = []
        for r in results:
            all_details.extend(r.get("details", []))
        ok = any(r.get("ok") for r in results)
        return {"ok": ok, "details": all_details, "total": sum(d.get("imported", 0) for d in all_details)}

    ext = path.suffix.lower()
    if ext == ".csv":
        result = import_csv_auto(path, db_path)
    elif ext in (".xlsx", ".xls"):
        result = import_excel(path, db_path)
    elif ext in (".db", ".sqlite", ".sqlite3"):
        result = import_sqlite_file(path, db_path)
    elif ext == ".sql":
        result = import_sql_file(path, db_path)
    else:
        return {"ok": False, "error": f"Formato nao suportado: {ext}", "details": []}

    if result.get("ok"):
        db.set_meta("atualizado_em", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), db_path)
    return result


def format_import_report(result: dict) -> str:
    if not result.get("ok"):
        return result.get("error", "Erro desconhecido")
    lines = ["Importacao concluida:\n"]
    for d in result.get("details", []):
        if d.get("ok"):
            lines.append(f"  • {d['table']}: {d['imported']} registos")
    total = result.get("total") or sum(d.get("imported", 0) for d in result.get("details", []))
    lines.append(f"\nTotal: {total} registos")
    return "\n".join(lines)
